"""Inspect all columns of BFG/Venom/Cooper/Apollo sheets."""
import openpyxl

wb = openpyxl.load_workbook('ADZ-RUGGEd-PRO_INVENTORY-SYSTEM-1.xlsx', data_only=True)

target_keywords = ['BFG', 'VENOM', 'COOPER', 'APOLLO']
for name in wb.sheetnames:
    if any(k in name.upper() for k in target_keywords):
        ws = wb[name]
        print(f"\n=== Sheet: {name!r} ===")
        # Show header row that has column labels
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            if any(v for v in row if v and str(v).strip()):
                print(f"  Row {i}: {[str(v)[:20] if v else '' for v in row[:12]]}")
        print("  --- first 10 data rows (all cols 0-11) ---")
        count = 0
        for row in ws.iter_rows(min_row=6, values_only=True):
            sku = row[0]
            if not sku or str(sku).strip() in ('', 'ITEM CODE') or str(sku).startswith('#'):
                continue
            print(f"  {[str(v)[:15] if v is not None else '' for v in row[:12]]}")
            count += 1
            if count >= 10:
                break
